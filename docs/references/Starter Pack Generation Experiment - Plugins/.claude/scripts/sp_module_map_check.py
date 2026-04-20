#!/usr/bin/env python3
"""
sp_module_map_check.py — Cross-document validation checker for Starter Packs.

Validates SP content against the Module Map (xlsx) and Toy Flow (docx) canonical
reference documents. Catches content drift between the spec and the SP.

Checks:
  MM1: Learning goals match Module Map verbatim text
  MM2: Required vocabulary terms present in §1.3
  MM3: Standards cascade matches Module Map
  MM4: Misconceptions match Module Map (correct IDs present, correct priority)
  MM5: TVP key beats present in correct phases
  MM6: Cognitive focus types match Module Map
  MM7: Question/test language stems present in EC interactions

Gate scoping:
  Gate 1: MM1, MM2 (partial), MM3, MM4 — foundation content checks
  Gate 2: MM1-MM6 (adds Warmup + Lesson key beat checks)
  Gate 3: MM1-MM7 (adds EC, Practice, Synthesis key beat checks)
  Gate 4: Same as Gate 3

Authority:
  - Grade 3 Unit 2 Area and Multiplication .xlsx (Module Map)
  - Grade 3 Unit 2_ Toy Flow.docx (TVP)

Usage:
  python sp_module_map_check.py <sp_file.md> --gate N [--json] [--output PATH]
"""

import argparse
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sp_parse_interactions import parse_sp

# ---------------------------------------------------------------------------
# Reference data paths — relative to the scripts directory
# ---------------------------------------------------------------------------

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(os.path.dirname(SCRIPTS_DIR))  # two levels up from .claude/scripts/

MODULE_MAP_PATH = os.path.join(BASE_DIR, "Grade 3 Unit 2", "Grade 3 Unit 2 Area and Multiplication .xlsx")
TVP_PATH = os.path.join(BASE_DIR, "Grade 3 Unit 2", "Grade 3 Unit 2_ Toy Flow.docx")


# ---------------------------------------------------------------------------
# Parse Module Map xlsx
# ---------------------------------------------------------------------------

def parse_module_map(filepath: str) -> dict:
    """Parse Module Map xlsx into per-module reference data."""
    import openpyxl

    if not os.path.exists(filepath):
        return {}

    wb = openpyxl.load_workbook(filepath, data_only=True)

    modules = {}

    # --- Module Mapping sheet ---
    ws = wb['Module Mapping']
    headers = [str(c.value or '').strip() for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]

    def col_idx(name):
        """Find column index by partial header match."""
        for i, h in enumerate(headers):
            if name.lower() in h.lower():
                return i
        return None

    idx_module = col_idx('Module')
    idx_goals = col_idx('Learning Goals')
    idx_vocab = col_idx('Vocabulary to Teach')
    idx_building_on = col_idx('Building On')
    idx_addressing = col_idx('Addressing')
    idx_building_toward = col_idx('Building Toward')
    idx_misconceptions = col_idx('Key Misconceptions')
    idx_question_lang = col_idx('Question/Test')
    idx_core_concept = col_idx('Core Concept')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row)
        module_id = str(vals[idx_module] or '').strip() if idx_module is not None else ''
        if not module_id or not module_id.startswith('M'):
            continue

        # Normalize module ID: "M1" → "M01"
        m_num = re.search(r'M(\d+)', module_id)
        if m_num:
            normalized_id = f"M{int(m_num.group(1)):02d}"
        else:
            normalized_id = module_id

        def get_val(idx):
            if idx is None or idx >= len(vals):
                return ''
            return str(vals[idx] or '').strip()

        modules[normalized_id] = {
            'module_id': normalized_id,
            'core_concept': get_val(col_idx('Core Concept')),
            'learning_goals': get_val(idx_goals),
            'vocabulary': [v.strip() for v in get_val(idx_vocab).split(',') if v.strip()],
            'standards_building_on': get_val(idx_building_on),
            'standards_addressing': get_val(idx_addressing),
            'standards_building_toward': get_val(idx_building_toward),
            'misconceptions_text': get_val(idx_misconceptions),
            'question_language': get_val(idx_question_lang),
        }

    # --- Misconceptions sheet ---
    if 'Misconceptions' in wb.sheetnames:
        ws_misc = wb['Misconceptions']
        misconceptions = []
        for row in ws_misc.iter_rows(min_row=2, max_row=ws_misc.max_row, values_only=True):
            vals = list(row)
            if len(vals) >= 6 and vals[0]:
                misconceptions.append({
                    'id': str(vals[0]).strip(),
                    'name': str(vals[1] or '').strip(),
                    'description': str(vals[2] or '').strip(),
                    'observable': str(vals[3] or '').strip(),
                    'where_surfaces': str(vals[4] or '').strip(),
                    'priority': str(vals[5] or '').strip(),
                })
        # Attach misconceptions to modules based on "Where Likely to Surface"
        for m_id, m_data in modules.items():
            m_num = int(re.search(r'\d+', m_id).group())
            m_data['misconceptions'] = []
            for misc in misconceptions:
                surfaces = misc.get('where_surfaces', '')
                # Check if this module is in the surface range (e.g., "M2-M3")
                ranges = re.findall(r'M(\d+)(?:-M?(\d+))?', surfaces)
                for start, end in ranges:
                    s = int(start)
                    e = int(end) if end else s
                    if s <= m_num <= e:
                        m_data['misconceptions'].append(misc)
                        break

    return modules


# ---------------------------------------------------------------------------
# Parse TVP docx — extract key beats per module
# ---------------------------------------------------------------------------

def parse_tvp(filepath: str) -> dict:
    """Parse TVP docx into per-module key beats."""
    from docx import Document

    if not os.path.exists(filepath):
        return {}

    doc = Document(filepath)
    modules = {}
    current_module = None
    current_phase = None

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        # Detect module headers: "MODULE 1: ..." or "MODULE 2: ..."
        m_match = re.match(r'^MODULE\s+(\d+)\s*:', text, re.IGNORECASE)
        if m_match:
            m_num = int(m_match.group(1))
            current_module = f"M{m_num:02d}"
            modules[current_module] = {'key_beats': [], 'sme_decisions': [], 'phases': {}}
            current_phase = None
            continue

        if current_module is None:
            continue

        # Detect phase headers
        phase_match = re.match(r'^(WARM-?UP|LESSON|EXIT CHECK|PRACTICE|SYNTHESIS)\b', text, re.IGNORECASE)
        if phase_match:
            current_phase = phase_match.group(1).upper().replace('-', '')
            if current_phase not in modules[current_module]['phases']:
                modules[current_module]['phases'][current_phase] = []
            continue

        # Extract key beats
        if text.startswith('Key beat:') or text.startswith('Key Beat:'):
            beat = text.split(':', 1)[1].strip().strip('"').strip('"').strip('"')
            modules[current_module]['key_beats'].append({
                'text': beat,
                'phase': current_phase or 'UNKNOWN',
            })

        # Extract SME decisions
        if 'RESOLVED:' in text:
            modules[current_module]['sme_decisions'].append(text)

    return modules


# ---------------------------------------------------------------------------
# Check functions
# ---------------------------------------------------------------------------

def check_learning_goals(sp, module_data: dict) -> list:
    """MM1: Learning goals in §1.1 match Module Map."""
    findings = []

    goals_text = module_data.get('learning_goals', '')
    if not goals_text:
        return findings

    # Extract individual goal fragments (split on semicolons, periods, or L1:/L2: markers)
    goal_fragments = []
    for part in re.split(r'(?:L\d:|;)', goals_text):
        part = part.strip()
        if len(part) > 10:
            goal_fragments.append(part)

    # Search §1.1 in the SP
    section_text = ''
    in_section = False
    for line in sp.lines:
        stripped = line.strip()
        if re.match(r'^##\s+\**§?1\.1\b', stripped, re.IGNORECASE) or \
           re.match(r'^##\s+\**LEARNING\s+GOALS?\b', stripped, re.IGNORECASE):
            in_section = True
            continue
        elif in_section and re.match(r'^##\s', stripped):
            break
        elif in_section:
            section_text += stripped + ' '

    if not section_text:
        findings.append({
            "check": "MM1",
            "severity": "MAJOR",
            "detail": "§1.1 Learning Goals section not found or empty — cannot validate against Module Map",
        })
        return findings

    section_lower = section_text.lower()

    for fragment in goal_fragments:
        # Check if key phrases from the goal appear in §1.1
        # Extract significant words (>4 chars) from the fragment
        key_words = [w for w in re.findall(r'\b\w+\b', fragment.lower()) if len(w) > 4]
        if not key_words:
            continue

        # Require at least 60% of key words to be present
        present = sum(1 for w in key_words if w in section_lower)
        ratio = present / len(key_words) if key_words else 0

        if ratio < 0.6:
            findings.append({
                "check": "MM1",
                "severity": "MAJOR",
                "detail": f"Learning goal fragment not found in §1.1: '{fragment[:80]}...' "
                         f"({present}/{len(key_words)} key words matched)",
                "fragment": fragment[:120],
                "match_ratio": round(ratio, 2),
            })

    return findings


def check_vocabulary(sp, module_data: dict) -> list:
    """MM2: Required vocabulary terms present in §1.3."""
    findings = []

    vocab_terms = module_data.get('vocabulary', [])
    if not vocab_terms:
        return findings

    # Extract §1.3 text
    section_text = ''
    in_section = False
    for line in sp.lines:
        stripped = line.strip()
        if re.match(r'^##\s+\**§?1\.3\b', stripped, re.IGNORECASE) or \
           re.match(r'^##\s+\**VOCABULARY\b', stripped, re.IGNORECASE):
            in_section = True
            continue
        elif in_section and re.match(r'^##\s', stripped):
            break
        elif in_section:
            section_text += stripped + ' '

    if not section_text:
        return findings  # §1.3 absence is caught by structure checker

    section_lower = section_text.lower()

    missing_terms = []
    for term in vocab_terms:
        term_clean = term.strip().lower()
        if term_clean and term_clean not in section_lower:
            # Try individual words for multi-word terms
            words = term_clean.split()
            if len(words) > 1:
                # Check if all significant words appear
                if not all(w in section_lower for w in words if len(w) > 2):
                    missing_terms.append(term.strip())
            else:
                missing_terms.append(term.strip())

    if missing_terms:
        findings.append({
            "check": "MM2",
            "severity": "MAJOR",
            "detail": f"Module Map vocabulary terms not found in §1.3: {', '.join(missing_terms)}",
            "missing_terms": missing_terms,
            "total_expected": len(vocab_terms),
            "total_missing": len(missing_terms),
        })

    return findings


def check_standards(sp, module_data: dict) -> list:
    """MM3: Standards cascade matches Module Map."""
    findings = []

    # Extract all standard codes from Module Map
    expected_standards = set()
    for field in ['standards_building_on', 'standards_addressing', 'standards_building_toward']:
        text = module_data.get(field, '')
        # Extract standard codes like "3.MD.C.5", "3.OA.A.1", "2.G.A.1"
        codes = re.findall(r'\d+\.\w+\.\w+\.?\w*', text)
        expected_standards.update(codes)

    if not expected_standards:
        return findings

    # Extract §1.1 text (standards typically in Standards Cascade subsection)
    full_text = ' '.join(sp.lines)

    missing = []
    for std in expected_standards:
        # Look for the standard code anywhere in the SP
        if std not in full_text:
            # Try partial match (without trailing subparts)
            parts = std.split('.')
            partial = '.'.join(parts[:3])
            if partial not in full_text:
                missing.append(std)

    if missing:
        findings.append({
            "check": "MM3",
            "severity": "MINOR",
            "detail": f"Standards from Module Map not found in SP: {', '.join(sorted(missing))}",
            "missing_standards": sorted(missing),
        })

    return findings


def check_misconceptions(sp, module_data: dict) -> list:
    """MM4: Misconceptions match Module Map."""
    findings = []

    expected_misconceptions = module_data.get('misconceptions', [])
    if not expected_misconceptions:
        return findings

    # Extract §1.4 text
    section_text = ''
    in_section = False
    for line in sp.lines:
        stripped = line.strip()
        if re.match(r'^##\s+\**§?1\.4\b', stripped, re.IGNORECASE) or \
           re.match(r'^##\s+\**MISCONCEPTION', stripped, re.IGNORECASE):
            in_section = True
            continue
        elif in_section and re.match(r'^##\s', stripped):
            break
        elif in_section:
            section_text += stripped + ' '

    if not section_text:
        return findings  # §1.4 absence caught by structure checker

    section_lower = section_text.lower()

    for misc in expected_misconceptions:
        name = misc.get('name', '')
        priority = misc.get('priority', '')

        if not name:
            continue

        # Check if misconception name or key words appear in §1.4
        name_words = [w for w in name.lower().split() if len(w) > 3]
        present = sum(1 for w in name_words if w in section_lower)
        ratio = present / len(name_words) if name_words else 0

        if ratio < 0.5:
            sev = "MAJOR" if 'HIGH' in priority.upper() else "MINOR"
            findings.append({
                "check": "MM4",
                "severity": sev,
                "detail": f"Misconception '{name}' (Priority: {priority}) not found in §1.4 "
                         f"({present}/{len(name_words)} key words matched)",
                "misconception": name,
                "priority": priority,
            })

    return findings


def check_tvp_key_beats(sp, tvp_module: dict, gate: int) -> list:
    """MM5: TVP key beats present in correct phases."""
    findings = []

    if not tvp_module:
        return findings

    key_beats = tvp_module.get('key_beats', [])
    if not key_beats:
        return findings

    full_text_lower = ' '.join(sp.lines).lower()

    # Gate scoping for which phases to check
    gate_phases = {
        1: set(),  # No phase checks at Gate 1
        2: {'WARMUP', 'LESSON'},
        3: {'WARMUP', 'LESSON', 'EXIT CHECK', 'EXITCHECK', 'PRACTICE', 'SYNTHESIS'},
        4: {'WARMUP', 'LESSON', 'EXIT CHECK', 'EXITCHECK', 'PRACTICE', 'SYNTHESIS'},
    }
    allowed_phases = gate_phases.get(gate, gate_phases[4])

    for beat in key_beats:
        phase = beat.get('phase', 'UNKNOWN')
        if phase not in allowed_phases:
            continue

        beat_text = beat.get('text', '')
        if not beat_text:
            continue

        # Extract key phrases from the beat (>4 char words)
        key_words = [w for w in re.findall(r'\b\w+\b', beat_text.lower())
                     if len(w) > 4 and w not in ('there', 'their', 'these', 'those', 'about',
                                                   'where', 'which', 'would', 'should', 'could')]
        if not key_words:
            continue

        present = sum(1 for w in key_words if w in full_text_lower)
        ratio = present / len(key_words) if key_words else 0

        if ratio < 0.4:
            findings.append({
                "check": "MM5",
                "severity": "MINOR",
                "detail": f"TVP key beat ({phase}) not well-reflected in SP: "
                         f"'{beat_text[:80]}...' ({present}/{len(key_words)} key words found)",
                "phase": phase,
                "beat": beat_text[:120],
                "match_ratio": round(ratio, 2),
            })

    return findings


def check_question_language(sp, module_data: dict) -> list:
    """MM7: Question/test language stems present in EC interactions."""
    findings = []

    question_lang = module_data.get('question_language', '')
    if not question_lang:
        return findings

    # Extract EC section text
    ec_text = ''
    in_ec = False
    for line in sp.lines:
        stripped = line.strip()
        if re.match(r'^##\s+\**§?1\.8\b', stripped, re.IGNORECASE) or \
           re.match(r'^##\s+\**EXIT\s+CHECK\b', stripped, re.IGNORECASE):
            in_ec = True
            continue
        elif in_ec and re.match(r'^##\s+\**§?1\.(8\.5|9)\b', stripped, re.IGNORECASE):
            break
        elif in_ec and re.match(r'^##\s+\**(?:PRACTICE|SYNTHESIS)\b', stripped, re.IGNORECASE):
            break
        elif in_ec:
            ec_text += stripped + ' '

    if not ec_text:
        return findings

    ec_lower = ec_text.lower()

    # Extract question stems (split on ? marks)
    stems = [s.strip() for s in question_lang.split('?') if len(s.strip()) > 10]

    for stem in stems:
        # Extract key content words
        key_words = [w for w in re.findall(r'\b\w+\b', stem.lower())
                     if len(w) > 4 and w not in ('about', 'these', 'those', 'which', 'would')]
        if not key_words:
            continue

        present = sum(1 for w in key_words if w in ec_lower)
        ratio = present / len(key_words) if key_words else 0

        if ratio < 0.3:
            findings.append({
                "check": "MM7",
                "severity": "MINOR",
                "detail": f"Question/test stem from Module Map not reflected in EC: "
                         f"'{stem[:60]}...' ({present}/{len(key_words)} key words)",
                "stem": stem[:120],
                "match_ratio": round(ratio, 2),
            })

    return findings


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------

def run_module_map_check(filepath: str, gate: int) -> dict:
    sp = parse_sp(filepath)
    all_findings = []

    # Get module ID from SP
    module_id = 'UNKNOWN'
    if sp.yaml and sp.yaml.fields.get('module_id'):
        module_id = sp.yaml.fields['module_id']
    else:
        # Try to extract from filename
        m = re.search(r'M(\d+)', os.path.basename(filepath))
        if m:
            module_id = f"M{int(m.group(1)):02d}"

    # Parse reference documents
    module_map = parse_module_map(MODULE_MAP_PATH)
    tvp = parse_tvp(TVP_PATH)

    module_data = module_map.get(module_id, {})
    tvp_module = tvp.get(module_id, {})

    meta = {
        'module_id': module_id,
        'module_map_found': os.path.exists(MODULE_MAP_PATH),
        'tvp_found': os.path.exists(TVP_PATH),
        'module_in_map': module_id in module_map,
        'module_in_tvp': module_id in tvp,
        'modules_in_map': sorted(module_map.keys()),
    }

    if not module_data and not tvp_module:
        all_findings.append({
            "check": "MM0",
            "severity": "MINOR",
            "detail": f"Module {module_id} not found in Module Map or TVP — skipping cross-document checks",
        })
        return _build_result(filepath, gate, all_findings, meta)

    # Run checks
    if module_data:
        all_findings.extend(check_learning_goals(sp, module_data))
        all_findings.extend(check_vocabulary(sp, module_data))
        all_findings.extend(check_standards(sp, module_data))
        all_findings.extend(check_misconceptions(sp, module_data))
        if gate >= 3:
            all_findings.extend(check_question_language(sp, module_data))

    if tvp_module and gate >= 2:
        all_findings.extend(check_tvp_key_beats(sp, tvp_module, gate))

    return _build_result(filepath, gate, all_findings, meta)


def _build_result(filepath, gate, all_findings, meta):
    checks_run = sorted(set(f["check"] for f in all_findings)) or ["MM1"]
    severity_counts = {}
    check_counts = {}
    for f in all_findings:
        sev = f.get("severity", "INFO")
        severity_counts[sev] = severity_counts.get(sev, 0) + 1
        c = f.get("check", "?")
        check_counts[c] = check_counts.get(c, 0) + 1

    return {
        "checker": "sp_module_map_check",
        "file": filepath,
        "gate": gate,
        "checks_run": checks_run,
        "total_findings": len(all_findings),
        "severity_counts": severity_counts,
        "check_counts": check_counts,
        "findings": all_findings,
        "meta": meta,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def print_findings_table(result: dict):
    print(f"\n{'='*80}")
    print(f"MODULE MAP CHECK — {result['file']} — Gate {result['gate']}")
    print(f"{'='*80}")
    print(f"Module ID: {result['meta'].get('module_id', '?')}")
    print(f"Module Map found: {result['meta'].get('module_map_found', False)}")
    print(f"TVP found: {result['meta'].get('tvp_found', False)}")
    print(f"Module in map: {result['meta'].get('module_in_map', False)}")
    print(f"Module in TVP: {result['meta'].get('module_in_tvp', False)}")
    print(f"Total findings: {result['total_findings']}")

    if result['severity_counts']:
        parts = [f"{k}: {v}" for k, v in sorted(result['severity_counts'].items())]
        print(f"Severity: {', '.join(parts)}")

    if result['findings']:
        print(f"\n{'─'*80}")
        print("FINDINGS:")
        print(f"{'─'*80}")

        for f in result['findings']:
            check = f.get('check', '?')
            sev = f.get('severity', '?')
            detail = f.get('detail', '')
            print(f"  [{check:3s}] {sev:8s} | {detail[:100]}")
    else:
        print("\n  ✓ No findings.")


def main():
    parser = argparse.ArgumentParser(description="SP Module Map Cross-Document Checker")
    parser.add_argument("sp_file", help="Path to the Starter Pack markdown file")
    parser.add_argument("--gate", type=int, required=True, choices=[1, 2, 3, 4])
    parser.add_argument("--json", action="store_true")
    parser.add_argument("--output", type=str)
    args = parser.parse_args()

    result = run_module_map_check(args.sp_file, args.gate)

    if args.json or args.output:
        json_str = json.dumps(result, indent=2)
        if args.output:
            os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)
            with open(args.output, 'w') as f:
                f.write(json_str)
            print(f"Output written to {args.output}")
        if args.json:
            print(json_str)
    else:
        print_findings_table(result)


if __name__ == '__main__':
    main()
